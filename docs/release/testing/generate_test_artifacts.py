#!/usr/bin/env python3
"""
Generate the staging validation test artifacts for PharmApp testers:
  - staging_validation_test_manual.pdf  (how-to guide, English)
  - staging_validation_test_cases.xlsx  (executable checklist, English)

Reproducible: re-run after editing TEST_CASES / META below.
    python -m pip install openpyxl fpdf2
    python docs/release/testing/generate_test_artifacts.py
"""

import os
from datetime import date

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT_DIR = os.path.dirname(os.path.abspath(__file__))
PDF_PATH = os.path.join(OUT_DIR, "staging_validation_test_manual.pdf")
XLSX_PATH = os.path.join(OUT_DIR, "staging_validation_test_cases.xlsx")

META = {
    "title": "PharmApp - Staging Validation Test Manual",
    "version": "1.0",
    "date": date.today().isoformat(),
    "app_url": "https://mediexchange-staging.web.app",
    "admin_url": "https://mediexchange-staging-admin.web.app",
    "admin_login": "admin@promoshake.net  /  Admin1234!",
    "project": "mediexchange-staging (isolated staging Firebase project - NOT production)",
}

# id, module, scenario, priority, preconditions, steps, data, expected
TEST_CASES = [
    ("TC-REG-01", "Registration", "S1", "P1",
     "App open, no account yet.",
     "Open app > choose Pharmacy > Register. Fill name, city Accra, phone, email, password, country Ghana. Submit WITHOUT a license number.",
     "email: pharmaA@promoshake.net; country: Ghana; city: Accra",
     "App re-prompts for a license number (LICENSE_REQUIRED). No usable account is created without a license."),
    ("TC-REG-02", "Registration", "S2", "P1",
     "TC-REG-01 license prompt is shown.",
     "Enter a valid license number and submit.",
     "license: GH-1234",
     "Account is created with status 'pending_verification'. Marketplace is not accessible yet."),
    ("TC-REG-03", "Registration", "S2", "P2",
     "App open.",
     "Register a Ghana pharmacy with an invalid license format.",
     "license: 1234 (wrong format)",
     "Rejected: license number does not match the required format (GH-####)."),
    ("TC-REG-04", "Registration", "-", "P2",
     "App open.",
     "Register a Cameroon pharmacy (country where no license is required).",
     "email: pharmaCM@promoshake.net; country: Cameroon; city: Douala",
     "Account created with status 'not_required'; trial subscription active immediately."),
    ("TC-AUTH-01", "Authentication", "-", "P1",
     "An account exists.",
     "Log out, then log back in with email + password.",
     "any created @promoshake.net account",
     "Login succeeds and lands on the correct dashboard."),
    ("TC-GATE-01", "Marketplace gate", "S7", "P1",
     "Pharmacy in status 'pending_verification'.",
     "Attempt to open the marketplace / create a medicine request before verification.",
     "-",
     "Action is blocked with a 'verification required' style message."),
    ("TC-ADM-01", "Admin - Auth", "S3", "P1",
     "Admin URL reachable.",
     "Log into the admin console.",
     "admin@promoshake.net / Admin1234!",
     "Admin dashboard loads."),
    ("TC-ADM-02", "Admin - License", "S3", "P1",
     "A pending Ghana pharmacy exists (TC-REG-02).",
     "Open 'License Reviews'. Confirm the pending pharmacy is listed (within country scope).",
     "-",
     "The pending pharmacy appears in the review list."),
    ("TC-ADM-03", "Admin - License", "S3", "P1",
     "TC-ADM-02.",
     "Click Verify on the pharmacy.",
     "-",
     "Status becomes 'verified'; trial subscription starts; pharmacy can now access the marketplace."),
    ("TC-ADM-04", "Admin - License", "-", "P2",
     "A pending pharmacy exists.",
     "Reject a license, providing a mandatory reason.",
     "reason: 'document unreadable'",
     "Status becomes 'rejected'; reason is recorded; pharmacy can submit a correction."),
    ("TC-ADM-05", "Admin - Country", "-", "P3",
     "Logged in as super_admin.",
     "Open Countries > edit Ghana license configuration and save.",
     "-",
     "Configuration saved successfully (no error)."),
    ("TC-PRO-01", "Profile", "-", "P2",
     "Verified pharmacy.",
     "Open profile > license status section.",
     "-",
     "Shows a 'Verified' badge."),
    ("TC-SUB-01", "Subscription", "-", "P2",
     "Verified pharmacy.",
     "Open the subscription screen.",
     "-",
     "Shows trial active with an end date (or 'trial_pending_license' before verification)."),
    ("TC-INV-01", "Inventory", "S4", "P1",
     "Verified pharmacy (seller).",
     "Add a medicine from the WHO essential database; set quantity and expiry.",
     "Paracetamol syrup; qty 50; future expiry",
     "Item appears in the pharmacy inventory."),
    ("TC-INV-02", "Inventory", "-", "P3",
     "Verified pharmacy.",
     "Add a medicine by scanning a barcode (EAN/UPC/QR).",
     "any scannable code",
     "Item added, or graceful 'unknown medicine' fallback."),
    ("TC-INV-03", "Inventory", "-", "P2",
     "Verified pharmacy.",
     "Add a custom medicine manually.",
     "custom name/dosage/form",
     "Item added to inventory."),
    ("TC-INV-04", "Inventory", "S5", "P2",
     "Verified pharmacy (buyer/requester).",
     "Add an exchange-currency item for barter.",
     "Ibuprofen 400mg; qty 60",
     "Item appears in the buyer inventory."),
    ("TC-WAL-01", "Wallet", "S4", "P1",
     "Verified pharmacy with an @promoshake.net email.",
     "Open Sandbox Testing screen > credit the wallet.",
     "amount: 100000",
     "Wallet 'available' balance increases. (Known: currency may show XAF - cosmetic.)"),
    ("TC-WAL-02", "Wallet", "-", "P2",
     "Wallet has a balance.",
     "View the wallet balance on the dashboard.",
     "-",
     "Balance is displayed correctly."),
    ("TC-MRP-01", "Medicine Request - Purchase", "S4", "P1",
     "2 verified Accra pharmacies; seller has inventory.",
     "Buyer: create a PURCHASE medicine request.",
     "Paracetamol syrup; qty 10",
     "Request created with status 'Open'."),
    ("TC-MRP-02", "Medicine Request - Purchase", "S4", "P1",
     "TC-MRP-01.",
     "Seller: submit a purchase offer (with a unit price) on the request.",
     "unit price: 50",
     "Offer created with status 'pending'."),
    ("TC-MRP-03", "Medicine Request - Purchase", "S4", "P1",
     "TC-MRP-02; buyer wallet funded.",
     "Buyer: accept the offer.",
     "-",
     "Buyer wallet debited by total; proposal 'accepted'; delivery 'pending'; request 'matched'; offer 'converted'. Seller stock unchanged at this step."),
    ("TC-MRP-04", "Medicine Request - Purchase", "-", "P2",
     "An open request with at least one pending offer.",
     "Buyer: cancel the request.",
     "-",
     "Request becomes 'cancelled'; pending offers become expired."),
    ("TC-MRE-01", "Medicine Request - Exchange", "S5", "P1",
     "Buyer holds item B (Ibuprofen); seller holds item A (Amoxicillin).",
     "Buyer: create an EXCHANGE medicine request for the seller's item.",
     "Amoxicillin 500mg; qty 5",
     "Request created with mode 'exchange'."),
    ("TC-MRE-02", "Medicine Request - Exchange", "S5", "P1",
     "TC-MRE-01.",
     "Seller: submit an exchange offer specifying the item wanted in return.",
     "wants: Ibuprofen 400mg; qty 5",
     "Offer created with an exchangeItem (no price)."),
    ("TC-MRE-03", "Medicine Request - Exchange", "S5", "P1",
     "TC-MRE-02.",
     "Buyer: accept; pick own Ibuprofen item in the inventory picker.",
     "-",
     "Only the buyer's item B is reserved (quantity held); seller item A unchanged at accept; NO wallet movement; proposal 'accepted' (exchange); delivery 'pending'."),
    ("TC-PAR-01", "Parity rules", "S6", "P2",
     "An open PURCHASE request exists.",
     "Seller: try to submit an EXCHANGE offer on the purchase request.",
     "-",
     "Rejected: offer type does not match the request mode."),
    ("TC-PAR-02", "Parity rules", "S6", "P2",
     "An open EXCHANGE request exists.",
     "Seller: try to submit a PURCHASE offer on the exchange request.",
     "-",
     "Rejected: offer type does not match the request mode."),
    ("TC-WDR-01", "Withdrawal", "S8", "P1",
     "Wallet funded.",
     "Create a withdrawal with a valid MTN Ghana number.",
     "msisdn: +233241234567; amount >= minimum and <= balance",
     "Withdrawal accepted (status 'processing'); wallet debited (moved to held)."),
    ("TC-WDR-02", "Withdrawal", "S8", "P2",
     "Wallet funded.",
     "Create a withdrawal with a wrong-operator number.",
     "msisdn: +233201234567 (Vodafone prefix)",
     "Rejected: MSISDN invalid for the selected provider."),
    ("TC-WDR-03", "Withdrawal", "S8", "P2",
     "Wallet funded.",
     "Create a withdrawal below the minimum amount.",
     "amount: very small (below GHS minimum)",
     "Rejected: amount below minimum withdrawal."),
    ("TC-NOT-01", "Notifications", "-", "P3",
     "An event happened (e.g. an offer was accepted).",
     "Check the notification bell / inbox.",
     "-",
     "A notification for the event appears in the inbox."),
]


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def build_xlsx():
    wb = Workbook()

    # ---- Sheet: Test Cases ----
    ws = wb.active
    ws.title = "Test Cases"
    headers = ["Test ID", "Module", "Scenario", "Priority", "Preconditions",
               "Test Steps", "Test Data", "Expected Result",
               "Actual Result", "Status", "Severity", "Tester", "Date", "Notes / Screenshot"]
    ws.append(headers)

    head_fill = PatternFill("solid", fgColor="1F4E78")
    head_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for tc in TEST_CASES:
        tid, mod, scen, prio, pre, steps, data, exp = tc
        ws.append([tid, mod, scen, prio, pre, steps, data, exp, "", "Not Run", "-", "", "", ""])

    widths = [12, 22, 9, 8, 32, 42, 28, 46, 30, 11, 11, 12, 12, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wrap_cols = [5, 6, 7, 8, 9, 14]
    last_row = ws.max_row
    for r in range(2, last_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=(c in wrap_cols))

    # Dropdowns
    status_dv = DataValidation(type="list", formula1='"Not Run,Pass,Fail,Blocked"', allow_blank=True)
    sev_dv = DataValidation(type="list", formula1='"-,Critical,Major,Minor"', allow_blank=True)
    ws.add_data_validation(status_dv)
    ws.add_data_validation(sev_dv)
    status_dv.add(f"J2:J{last_row}")
    sev_dv.add(f"K2:K{last_row}")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:N{last_row}"

    # ---- Sheet: Environment & Access ----
    env = wb.create_sheet("Environment & Access")
    env_rows = [
        ("PharmApp - Staging Validation", ""),
        ("", ""),
        ("Document version", META["version"]),
        ("Generated", META["date"]),
        ("Environment", META["project"]),
        ("", ""),
        ("Pharmacy app URL", META["app_url"]),
        ("Admin console URL", META["admin_url"]),
        ("Admin login", META["admin_login"]),
        ("", ""),
        ("IMPORTANT - test email rule", "Register every test pharmacy with an email ending in @promoshake.net"),
        ("", "(otherwise wallet funding via the Sandbox screen is rejected: NOT_TEST_ACCOUNT)"),
        ("Ghana license format", "GH-#### (e.g. GH-1234)"),
        ("Test city", "Accra (Ghana) for marketplace tests - both pharmacies must be in the same city"),
        ("MTN Ghana MSISDN", "+23324XXXXXXX (valid) ; +23320XXXXXXX = wrong operator (should be rejected)"),
        ("", ""),
        ("Known non-blocking issues", ""),
        ("Registration snackbar", "A 'Registration failed' message may appear EVEN WHEN the account is created. Verify the account exists before marking Fail."),
        ("Wallet currency", "A Ghana wallet may display XAF instead of GHS (cosmetic)."),
        ("", ""),
        ("How to report", "Fill columns Actual Result / Status / Severity / Tester / Date / Notes. Attach screenshots and reference the file name in Notes."),
        ("After testing", "Save and return this workbook; results will be uploaded and analysed."),
    ]
    for row in env_rows:
        env.append(list(row))
    env.column_dimensions["A"].width = 30
    env.column_dimensions["B"].width = 95
    env["A1"].font = Font(bold=True, size=14, color="1F4E78")
    bold_red_labels = {"IMPORTANT - test email rule", "Known non-blocking issues",
                       "How to report", "After testing"}
    for r in range(1, env.max_row + 1):
        if env.cell(row=r, column=1).value in bold_red_labels:
            env.cell(row=r, column=1).font = Font(bold=True, color="C00000")
        env.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")

    # ---- Sheet: Summary ----
    summ = wb.create_sheet("Summary")
    summ["A1"] = "Result Summary"
    summ["A1"].font = Font(bold=True, size=14, color="1F4E78")
    rng = f"'Test Cases'!J2:J{last_row}"
    summ.append([])
    summ.append(["Total test cases", len(TEST_CASES)])
    summ.append(["Pass", f'=COUNTIF({rng},"Pass")'])
    summ.append(["Fail", f'=COUNTIF({rng},"Fail")'])
    summ.append(["Blocked", f'=COUNTIF({rng},"Blocked")'])
    summ.append(["Not Run", f'=COUNTIF({rng},"Not Run")'])
    summ.column_dimensions["A"].width = 22
    summ.column_dimensions["B"].width = 14
    for r in range(3, 8):
        summ.cell(row=r, column=1).font = Font(bold=True)

    wb.save(XLSX_PATH)
    print(f"wrote {XLSX_PATH}  ({len(TEST_CASES)} test cases)")


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

def _t(s: str) -> str:
    """Make text latin-1 safe for the core PDF font."""
    return (s.replace("’", "'").replace("‘", "'")
             .replace("“", '"').replace("”", '"')
             .replace("–", "-").replace("—", "-")
             .replace("→", "->").encode("latin-1", "replace").decode("latin-1"))


class ManualPDF(FPDF):
    def header(self):
        if self.page_no() == 1:
            return
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(130)
        self.cell(0, 8, _t(META["title"]), align="L")
        self.cell(0, 8, f"v{META['version']}", align="R", new_x="LMARGIN", new_y="NEXT")
        self.set_text_color(0)

    def footer(self):
        self.set_y(-12)
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(130)
        self.cell(0, 8, f"Page {self.page_no()}", align="C")
        self.set_text_color(0)

    def h1(self, text):
        self.ln(2)
        self.set_font("Helvetica", "B", 14)
        self.set_text_color(31, 78, 120)
        self.multi_cell(0, 8, _t(text), new_x="LMARGIN", new_y="NEXT")
        self.set_text_color(0)
        self.ln(1)

    def h2(self, text):
        self.ln(1)
        self.set_font("Helvetica", "B", 11)
        self.multi_cell(0, 6, _t(text), new_x="LMARGIN", new_y="NEXT")

    def body(self, text):
        self.set_font("Helvetica", "", 10)
        self.multi_cell(0, 5.2, _t(text), new_x="LMARGIN", new_y="NEXT")

    def bullet(self, text):
        self.set_font("Helvetica", "", 10)
        x0 = self.get_x()
        self.multi_cell(0, 5.2, _t("  - " + text), new_x="LMARGIN", new_y="NEXT")
        self.set_x(x0)


def build_pdf():
    pdf = ManualPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Title block
    pdf.ln(20)
    pdf.set_font("Helvetica", "B", 22)
    pdf.set_text_color(31, 78, 120)
    pdf.multi_cell(0, 11, _t("PharmApp"), new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "B", 16)
    pdf.set_text_color(0)
    pdf.multi_cell(0, 9, _t("Staging Validation - Test Manual"), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)
    pdf.set_font("Helvetica", "", 11)
    pdf.multi_cell(0, 6, _t(f"Version {META['version']}   |   {META['date']}"), new_x="LMARGIN", new_y="NEXT")
    pdf.multi_cell(0, 6, _t("Environment: " + META["project"]), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(6)
    pdf.set_draw_color(31, 78, 120)
    pdf.set_line_width(0.6)
    pdf.line(pdf.l_margin, pdf.get_y(), pdf.w - pdf.r_margin, pdf.get_y())
    pdf.ln(8)

    pdf.h1("1. Purpose & scope")
    pdf.body("This manual guides testers through a full end-to-end validation of PharmApp on the isolated STAGING environment. Staging is a separate Firebase project; it does NOT affect production or real customer data. The goal is to validate the new features (pharmacy license workflow, trial subscription, medicine requests in purchase and exchange modes, mobile-money withdrawal) before they are promoted to production.")

    pdf.h1("2. Test environment & access")
    pdf.bullet("Pharmacy app: " + META["app_url"])
    pdf.bullet("Admin console: " + META["admin_url"])
    pdf.bullet("Admin login: " + META["admin_login"])
    pdf.bullet("Use any modern desktop browser (Chrome recommended).")

    pdf.h1("3. Test data conventions")
    pdf.bullet("Register EVERY test pharmacy with an email ending in @promoshake.net. This is required - otherwise wallet funding via the Sandbox screen is rejected (NOT_TEST_ACCOUNT).")
    pdf.bullet("Ghana license number format: GH-#### (for example GH-1234).")
    pdf.bullet("For marketplace tests, both pharmacies must be in the SAME city (use Accra, Ghana).")
    pdf.bullet("Valid MTN Ghana number: +23324XXXXXXX. A Vodafone-prefix number (+23320XXXXXXX) must be rejected.")
    pdf.bullet("Cameroon is a country where no license is required (useful for the no-license path).")

    pdf.h1("4. Known non-blocking issues")
    pdf.body("These are known cosmetic issues. Do NOT mark a test as Fail because of them - note them instead.")
    pdf.bullet("Registration: a 'Registration failed' message may appear EVEN WHEN the account was actually created. Verify the account exists (try logging in) before deciding.")
    pdf.bullet("Wallet currency: a Ghana wallet may show XAF instead of GHS. This is cosmetic.")

    pdf.h1("5. How to run a test and report results")
    pdf.bullet("Open the test cases workbook (staging_validation_test_cases.xlsx).")
    pdf.bullet("Execute each test case in order (some depend on previous ones - see Preconditions).")
    pdf.bullet("Record: Actual Result, Status (Pass / Fail / Blocked), Severity if failed, your name (Tester), the Date, and Notes.")
    pdf.bullet("Capture a screenshot for any failure and reference its file name in the Notes column.")
    pdf.bullet("When finished, save and return the workbook. Results will be uploaded and analysed.")

    pdf.h1("6. Recommended end-to-end flow")
    pdf.body("The test cases follow this storyline:")
    pdf.bullet("Register a Ghana pharmacy (license workflow) and a Cameroon pharmacy (no license).")
    pdf.bullet("As admin, verify the Ghana pharmacy license.")
    pdf.bullet("Add inventory and fund the wallet (Sandbox screen).")
    pdf.bullet("Create and complete a PURCHASE medicine request between two pharmacies.")
    pdf.bullet("Create and complete an EXCHANGE (barter) medicine request.")
    pdf.bullet("Verify cross-mode rejection rules.")
    pdf.bullet("Perform a wallet withdrawal (valid and invalid cases).")

    pdf.add_page()
    pdf.h1("7. Test cases (overview)")
    pdf.body("Full steps and expected results are in the Excel workbook. Summary below:")
    pdf.ln(1)

    # Table header
    col_w = [22, 40, 18, 14, 0]  # last auto (title)
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(31, 78, 120)
    pdf.set_text_color(255)
    pdf.cell(col_w[0], 7, "Test ID", border=1, fill=True)
    pdf.cell(col_w[1], 7, "Module", border=1, fill=True)
    pdf.cell(col_w[2], 7, "Scenario", border=1, fill=True)
    pdf.cell(col_w[3], 7, "Priority", border=1, fill=True, new_x="LMARGIN", new_y="NEXT")
    pdf.set_text_color(0)
    pdf.set_font("Helvetica", "", 9)
    fill = False
    for tid, mod, scen, prio, *_ in TEST_CASES:
        pdf.set_fill_color(238, 242, 248)
        pdf.cell(col_w[0], 6, _t(tid), border=1, fill=fill)
        pdf.cell(col_w[1], 6, _t(mod), border=1, fill=fill)
        pdf.cell(col_w[2], 6, _t(scen), border=1, fill=fill)
        pdf.cell(col_w[3], 6, _t(prio), border=1, fill=fill, new_x="LMARGIN", new_y="NEXT")
        fill = not fill

    pdf.ln(4)
    pdf.h2("Glossary")
    for term, desc in [
        ("pending_verification", "pharmacy registered but license not yet approved; cannot use the marketplace."),
        ("verified", "license approved by an admin; full marketplace access; trial starts."),
        ("medicine request", "a pharmacy asks for a medicine; others make offers (purchase or exchange)."),
        ("purchase offer", "an offer with a price; accepting it debits the buyer wallet."),
        ("exchange offer", "a barter offer; the seller asks for an item in return; no money moves."),
        ("withdrawal", "moving wallet funds out via mobile money (MTN Ghana on staging)."),
    ]:
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(40, 5.2, _t(term))
        pdf.set_font("Helvetica", "", 9)
        pdf.multi_cell(0, 5.2, _t(desc), new_x="LMARGIN", new_y="NEXT")

    pdf.output(PDF_PATH)
    print(f"wrote {PDF_PATH}")


if __name__ == "__main__":
    build_xlsx()
    build_pdf()
